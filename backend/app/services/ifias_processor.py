"""
IFIAS Processor — Core Orchestrator
Single class that runs the entire pipeline for one Excel billing file.

Flow:
    IfiasProcessor.run(excel_path, batch_id)
      1. Parse Excel (SL NO order)
      2. Load cache
      3. For each LR: cache → IMAP → OneDrive fallback → validate → extract
      4. Write Excel ONCE  →  *_processed.xlsx
      5. Generate report.json + report.txt
      6. Save cache
"""

import hashlib
import json
import logging
import os
import tempfile
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import PatternFill

from app.services.excel_parser_service import parse_invoice_excel, InvoiceLineItem
from app.services.email_intelligence_service import EmailIntelligenceService
from app.services.browser_email_service import BrowserEmailService
from app.services.satisfaction_slip_parser import SatisfactionSlipParser, pdf_contains_lr

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Status constants — exactly what gets written into the STATUS Excel column
# ---------------------------------------------------------------------------

STATUS_SUCCESS = "SUCCESS"
STATUS_MANUAL  = "REQUIRES MANUAL INSPECTION"
STATUS_ERROR   = "ERROR"

# Cell highlight colours
FILL_SUCCESS = PatternFill("solid", fgColor="C6EFCE")   # green
FILL_MANUAL  = PatternFill("solid", fgColor="FFEB9C")   # yellow
FILL_ERROR   = PatternFill("solid", fgColor="FFC7CE")   # red

# OneDrive fallback path
ONEDRIVE_WATCH_PATH = os.getenv(
    "ONEDRIVE_WATCH_PATH",
    os.path.join(os.path.expanduser("~"), "OneDrive"),
)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class LRResult:
    """Holds the processing outcome for a single LR row."""
    sl_no: Optional[int]
    lr_number: str
    excel_row_index: int
    truck_type: Optional[str] = None
    detention_days: Optional[int] = None
    status: str = STATUS_MANUAL
    error_detail: Optional[str] = None
    from_cache: bool = False
    source: Optional[str] = None    # CACHE | IMAP | BROWSER | ONEDRIVE


@dataclass
class ProcessorResult:
    """Returned by IfiasProcessor.run() after the full loop completes."""
    batch_id: Optional[int]
    excel_path: str
    processed_excel_path: str
    report_json_path: str
    report_txt_path: str
    total: int
    successful: int
    manual_required: int
    errors: int
    lr_results: List[LRResult]
    started_at: datetime
    completed_at: datetime


# ---------------------------------------------------------------------------
# Main orchestrator
# ---------------------------------------------------------------------------

class IfiasProcessor:
    """
    Core IFIAS orchestrator.

    Usage:
        processor = IfiasProcessor()
        result = processor.run(excel_path="/path/to/billing.xlsx", batch_id=1)
    """

    CACHE_DIR  = Path(os.getenv("IFIAS_CACHE_DIR",  str(Path.home() / ".ifias_cache")))
    OUTPUT_DIR = Path(os.getenv("IFIAS_OUTPUT_DIR", str(Path.home() / "ifias_output")))

    def __init__(self):
        self.CACHE_DIR.mkdir(parents=True, exist_ok=True)
        self.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------

    def run(self, excel_path: str, batch_id: Optional[int] = None) -> ProcessorResult:
        """
        Main entry point. Runs the full pipeline for one Excel billing file.

        Args:
            excel_path: Absolute path to the billing Excel file.
            batch_id:   DB ProcessingBatch.id (used for logging / report).

        Returns:
            ProcessorResult with paths to processed Excel and reports.
        """
        started_at = datetime.utcnow()
        logger.info(f"[IFIAS] Starting | file={excel_path}  batch_id={batch_id}")

        # STEP 1 — Parse Excel
        parse_result = parse_invoice_excel(excel_path)
        if not parse_result.line_items:
            raise ValueError(f"No LR rows found in: {excel_path}")

        logger.info(
            f"[IFIAS] Parsed {len(parse_result.line_items)} LRs "
            f"from sheet '{parse_result.sheet_name}'"
        )

        # Load persistent cache
        cache = self._load_cache(excel_path)

        # STEP 2 — Process each LR (NEVER stop loop)
        lr_results: List[LRResult] = []

        email_svc = EmailIntelligenceService()
        email_connected = email_svc.connect()
        if not email_connected:
            logger.warning("[IFIAS] IMAP connect failed — will try browser automation fallback")

        browser_svc = BrowserEmailService()
        browser_connected = browser_svc.connect()
        if not browser_connected:
            logger.info("[IFIAS] Browser fallback not available (credentials not set or Playwright not installed)")

        try:
            for item in parse_result.line_items:
                result = self._process_lr(
                    item, cache,
                    email_svc if email_connected else None,
                    browser_svc if browser_connected else None,
                )
                lr_results.append(result)

                # Update in-memory cache for successful results
                if result.status == STATUS_SUCCESS and not result.from_cache:
                    cache[self._cache_key(item)] = {
                        "truck_type":    result.truck_type,
                        "detention_days": result.detention_days,
                        "status":        STATUS_SUCCESS,
                        "source":        result.source,
                        "cached_at":     datetime.utcnow().isoformat(),
                    }
        finally:
            if email_connected:
                email_svc.disconnect()
            browser_svc.disconnect()

        # STEP 3 — Write Excel ONCE (after loop)
        stem = Path(excel_path).stem
        processed_path = str(self.OUTPUT_DIR / f"{stem}_processed.xlsx")
        self._write_excel(excel_path, lr_results, processed_path)

        # STEP 4 — Generate report
        from app.services.ifias_report import IfiasReportBuilder
        report = IfiasReportBuilder.build(
            batch_id=batch_id,
            excel_path=excel_path,
            lr_results=lr_results,
            started_at=started_at,
        )
        report_json_path = str(self.OUTPUT_DIR / f"{stem}_report.json")
        report_txt_path  = str(self.OUTPUT_DIR / f"{stem}_report.txt")
        IfiasReportBuilder.save(report, report_json_path, report_txt_path)

        # STEP 5 — Persist cache
        self._save_cache(excel_path, cache)

        completed_at = datetime.utcnow()
        logger.info(
            f"[IFIAS] Done | total={report['total']}  success={report['successful']}  "
            f"manual={report['manual_required']}  error={report['errors']}"
        )

        return ProcessorResult(
            batch_id=batch_id,
            excel_path=excel_path,
            processed_excel_path=processed_path,
            report_json_path=report_json_path,
            report_txt_path=report_txt_path,
            total=report["total"],
            successful=report["successful"],
            manual_required=report["manual_required"],
            errors=report["errors"],
            lr_results=lr_results,
            started_at=started_at,
            completed_at=completed_at,
        )

    # ------------------------------------------------------------------
    # Per-LR processing
    # ------------------------------------------------------------------

    def _process_lr(
        self,
        item: InvoiceLineItem,
        cache: Dict,
        email_svc: Optional[EmailIntelligenceService],
        browser_svc: Optional[BrowserEmailService] = None,
    ) -> LRResult:
        """
        Thin wrapper: runs _do_process_lr with a 20-second per-LR timeout.
        Never raises.
        """
        lr = item.lr_number
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(
                self._do_process_lr, item, cache, email_svc, browser_svc
            )
            try:
                return future.result(timeout=20)
            except FuturesTimeoutError:
                logger.warning(f"[{lr}] Processing timeout (20s) → ERROR")
                return LRResult(
                    sl_no=item.sl_no,
                    lr_number=lr,
                    excel_row_index=item.excel_row_index,
                    status=STATUS_ERROR,
                    error_detail="Processing timeout exceeded 20 seconds",
                )
            except Exception as exc:
                logger.error(f"[{lr}] Unexpected outer error: {exc}", exc_info=True)
                return LRResult(
                    sl_no=item.sl_no,
                    lr_number=lr,
                    excel_row_index=item.excel_row_index,
                    status=STATUS_ERROR,
                    error_detail=str(exc),
                )

    def _do_process_lr(
        self,
        item: InvoiceLineItem,
        cache: Dict,
        email_svc: Optional[EmailIntelligenceService],
        browser_svc: Optional[BrowserEmailService] = None,
    ) -> LRResult:
        """
        Core per-LR logic.
        Search order: CACHE → IMAP → BROWSER → ONEDRIVE
        pdf_contains_lr() is validated per-source before moving to the next.
        """
        lr = item.lr_number
        result = LRResult(
            sl_no=item.sl_no,
            lr_number=lr,
            excel_row_index=item.excel_row_index,
        )

        try:
            # 1. Cache check
            cache_key = self._cache_key(item)
            if cache_key in cache:
                cached = cache[cache_key]
                result.truck_type     = cached.get("truck_type")
                result.detention_days = cached.get("detention_days")
                result.status         = STATUS_SUCCESS
                result.source         = "CACHE"
                result.from_cache     = True
                logger.info(f"[{lr}] Cache hit")
                return result

            pdf_path: Optional[str] = None

            # 2. Try IMAP (validate per-source — if wrong PDF, continue to next source)
            if email_svc:
                candidate = self._fetch_from_email(lr, email_svc)
                if candidate and pdf_contains_lr(candidate, lr):
                    pdf_path = candidate
                    result.source = "IMAP"
                    logger.info(f"[{lr}] PDF validated via IMAP")
                elif candidate:
                    logger.warning(f"[{lr}] IMAP PDF does not contain LR — trying next source")

            # 3. Try Browser (if IMAP missed or returned wrong PDF)
            if not pdf_path and browser_svc:
                logger.info(f"[{lr}] Trying browser automation")
                candidate = self._fetch_from_browser(lr, browser_svc)
                if candidate and pdf_contains_lr(candidate, lr):
                    pdf_path = candidate
                    result.source = "BROWSER"
                    logger.info(f"[{lr}] PDF validated via Browser")
                elif candidate:
                    logger.warning(f"[{lr}] Browser PDF does not contain LR — trying OneDrive")

            # 4. Try OneDrive local folder (limited depth)
            if not pdf_path:
                candidate = self._fetch_from_onedrive(lr)
                if candidate and pdf_contains_lr(candidate, lr):
                    pdf_path = candidate
                    result.source = "ONEDRIVE"
                    logger.info(f"[{lr}] PDF validated via OneDrive")
                elif candidate:
                    logger.warning(f"[{lr}] OneDrive PDF does not contain LR")

            if not pdf_path:
                result.status       = STATUS_MANUAL
                result.error_detail = "No valid satisfaction slip PDF found (IMAP + Browser + OneDrive searched)"
                logger.info(f"[{lr}] No PDF found → MANUAL")
                return result

            # 5. Extract (truck_type + detention_days)
            parser = SatisfactionSlipParser()
            slip = parser.parse(pdf_path)

            if slip is None or (slip.truck_type is None and slip.detention_days is None):
                result.status       = STATUS_ERROR
                result.error_detail = "OCR extraction returned no usable data"
                logger.warning(f"[{lr}] OCR failed → ERROR")
                return result

            result.truck_type     = slip.truck_type
            result.detention_days = slip.detention_days
            result.status         = STATUS_SUCCESS
            logger.info(
                f"[{lr}] SUCCESS | source={result.source}  "
                f"truck_type={slip.truck_type}  detention={slip.detention_days}"
            )

        except Exception as exc:
            result.status       = STATUS_ERROR
            result.error_detail = str(exc)
            logger.error(f"[{lr}] Unexpected error: {exc}", exc_info=True)

        return result

    # ------------------------------------------------------------------
    # Email + OneDrive helpers
    # ------------------------------------------------------------------

    def _fetch_from_email(
        self, lr: str, email_svc: EmailIntelligenceService
    ) -> Optional[str]:
        """Search Outlook IMAP for a PDF attachment matching this LR. Returns local path or None."""
        try:
            best = email_svc.get_best_match(lr)
            if not best or not best.has_pdf_attachment:
                logger.info(f"[{lr}] No email match")
                return None

            tmp_dir  = tempfile.mkdtemp()
            safe_lr  = lr.replace("/", "_").replace(" ", "_")
            tmp_pdf  = os.path.join(tmp_dir, f"slip_{safe_lr}.pdf")
            dl = email_svc.download_attachment(best.message_id, best.folder, tmp_pdf)

            if dl.success:
                logger.info(f"[{lr}] Email PDF downloaded: {dl.local_path}")
                return dl.local_path

            logger.warning(f"[{lr}] Email download failed: {dl.error}")
        except Exception as exc:
            logger.warning(f"[{lr}] Email search error: {exc}")
        return None

    def _fetch_from_browser(
        self, lr: str, browser_svc: BrowserEmailService
    ) -> Optional[str]:
        """
        Use browser automation to find and download a PDF from Outlook Web.
        Returns local path to downloaded PDF, or None.
        """
        try:
            pdf_path = browser_svc.get_pdf_for_lr(lr)
            if pdf_path:
                logger.info(f"[{lr}] Browser PDF downloaded: {pdf_path}")
            return pdf_path
        except Exception as exc:
            logger.warning(f"[{lr}] Browser search error: {exc}")
        return None

    @staticmethod
    def _fetch_from_onedrive(lr: str, max_depth: int = 3) -> Optional[str]:
        """
        Walk the local OneDrive sync folder up to max_depth levels deep,
        searching for a PDF whose filename contains the LR number.
        Validation (pdf_contains_lr) is performed by the caller.
        Returns the first filename-matched path, or None.
        """
        if not os.path.isdir(ONEDRIVE_WATCH_PATH):
            return None

        normalized_lr = lr.replace(" ", "").upper()
        base_depth = ONEDRIVE_WATCH_PATH.rstrip(os.sep).count(os.sep)
        logger.info(f"[{lr}] Searching OneDrive (depth≤{max_depth}): {ONEDRIVE_WATCH_PATH}")

        for root, dirs, files in os.walk(ONEDRIVE_WATCH_PATH):
            current_depth = root.count(os.sep) - base_depth
            if current_depth >= max_depth:
                dirs[:] = []  # prune — do not recurse deeper

            for fname in files:
                if not fname.lower().endswith(".pdf"):
                    continue
                if normalized_lr in fname.replace(" ", "").upper():
                    full_path = os.path.join(root, fname)
                    logger.info(f"[{lr}] OneDrive candidate: {full_path}")
                    return full_path

        logger.info(f"[{lr}] Not found in OneDrive (depth≤{max_depth})")
        return None

    # ------------------------------------------------------------------
    # Excel writeback — called ONCE after the loop
    # ------------------------------------------------------------------

    def _write_excel(
        self,
        source_path: str,
        results: List[LRResult],
        output_path: str,
    ) -> None:
        """
        Opens source Excel, updates truck_type / detention_days / STATUS columns,
        saves to output_path. The original file is never modified.
        """
        wb = openpyxl.load_workbook(source_path)
        ws = self._detect_sheet(wb)

        header_row_idx, col_map = self._find_headers(ws)

        # Add STATUS column if missing
        if "status" not in col_map:
            new_col = ws.max_column + 1
            ws.cell(row=header_row_idx, column=new_col, value="STATUS")
            col_map["status"] = new_col

        # Add SOURCE column if missing
        if "source" not in col_map:
            new_col = ws.max_column + 1
            ws.cell(row=header_row_idx, column=new_col, value="SOURCE")
            col_map["source"] = new_col

        truck_type_col  = col_map.get("truck_type")
        detention_col   = col_map.get("detention_days")
        status_col      = col_map["status"]
        source_col      = col_map["source"]

        # Build lookup: excel_row_index → LRResult
        row_map: Dict[int, LRResult] = {r.excel_row_index: r for r in results}

        for row_idx, lr_result in row_map.items():
            if truck_type_col and lr_result.truck_type:
                ws.cell(row=row_idx, column=truck_type_col, value=lr_result.truck_type)
            if detention_col and lr_result.detention_days is not None:
                ws.cell(row=row_idx, column=detention_col, value=lr_result.detention_days)

            status_cell = ws.cell(row=row_idx, column=status_col, value=lr_result.status)
            if lr_result.status == STATUS_SUCCESS:
                status_cell.fill = FILL_SUCCESS
            elif lr_result.status == STATUS_MANUAL:
                status_cell.fill = FILL_MANUAL
            else:
                status_cell.fill = FILL_ERROR

            if lr_result.source:
                ws.cell(row=row_idx, column=source_col, value=lr_result.source)

        wb.save(output_path)
        logger.info(f"[IFIAS] Processed Excel saved → {output_path}")

    # ------------------------------------------------------------------
    # Cache helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _cache_key(item: InvoiceLineItem) -> str:
        """Deterministic cache key: hash(lr | shipment_no | truck_no)."""
        raw = f"{item.lr_number}|{item.shipment_no or ''}|{item.truck_number or ''}"
        return hashlib.sha1(raw.encode()).hexdigest()[:16]

    def _cache_path(self, excel_path: str) -> Path:
        """Each Excel file gets its own cache file keyed by its path hash."""
        stem = hashlib.sha1(excel_path.encode()).hexdigest()[:12]
        return self.CACHE_DIR / f"cache_{stem}.json"

    def _load_cache(self, excel_path: str) -> Dict:
        path = self._cache_path(excel_path)
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                logger.warning(f"[IFIAS] Cache corrupt, starting fresh: {path}")
        return {}

    def _save_cache(self, excel_path: str, cache: Dict) -> None:
        path = self._cache_path(excel_path)
        path.write_text(json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")
        logger.info(f"[IFIAS] Cache saved → {path}")

    # ------------------------------------------------------------------
    # Sheet / header detection helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _detect_sheet(wb: openpyxl.Workbook):
        """Return the first non-hidden sheet."""
        for name in wb.sheetnames:
            ws = wb[name]
            if ws.sheet_state != "hidden":
                return ws
        return wb.active

    # Column aliases — same as excel_parser_service for consistency
    _HEADER_ALIASES = {
        "truck_type":    ["DESC MEANS OF TRPT", "TRUCK TYPE", "TRPT TYPE", "VEHICLE TYPE", "DESC MEANS TRPT"],
        "detention_days": ["DETENTN DAYS", "DETENTION DAYS", "DET DAYS", "DETENTION"],
        "lr_number":     ["LR NO", "LR_NO", "LR NUMBER", "LR.NO"],
        "status":        ["STATUS"],
        "source":        ["SOURCE"],
    }

    def _find_headers(self, ws) -> tuple:
        """
        Scan rows 1-30 for the header row (identified by containing an LR column).
        Returns (header_row_idx, col_map) where col_map maps canonical keys to 1-based col indices.
        """
        for row in ws.iter_rows(min_row=1, max_row=30):
            vals = [str(c.value or "").strip().upper() for c in row]
            if any("LR" in v for v in vals):
                col_map: Dict[str, int] = {}
                for col_idx, val in enumerate(vals, start=1):
                    for key, aliases in self._HEADER_ALIASES.items():
                        if val in [a.upper() for a in aliases]:
                            col_map[key] = col_idx
                return row[0].row, col_map
        return 1, {}
