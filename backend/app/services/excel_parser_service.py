"""
Excel Parser Service — IFIAS Phase 1
Reads freight billing Excel files and extracts LR line items.

Usage:
    from app.services.excel_parser_service import parse_invoice_excel
    result = parse_invoice_excel("/path/to/SENTHIL TRANSPORTS MAR'26-02.xlsx")
"""

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class InvoiceLineItem:
    lr_number: str
    truck_number: Optional[str]
    sat_slip_no: Optional[str]
    detention_days: Optional[int]
    truck_type: Optional[str]
    shipment_no: Optional[str]
    service_po: Optional[str]
    entry_sheet_no: Optional[str]
    region: Optional[str]
    from_location: Optional[str]
    to_location: Optional[str]
    total_units: Optional[float]
    total_wt: Optional[float]
    shortage: Optional[float]
    detention_charge: Optional[float]
    master_rate: Optional[float]
    payable: Optional[float]
    remarks: Optional[str]
    needs_ocr: bool
    needs_ocr_verify: bool
    row_number: int


@dataclass
class InvoiceParseResult:
    transporter_name: str
    billing_period: str
    client_name: str
    total_rows: int
    rows_needing_ocr: int
    line_items: List[InvoiceLineItem]
    file_path: str
    parsed_at: datetime
    sheet_name: str
    errors: List[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Column name → canonical key mapping (case-insensitive fuzzy)
# ---------------------------------------------------------------------------

COLUMN_MAP = {
    "lr_number":       ["LR NO", "LR_NO", "LR NUMBER", "LR.NO"],
    "truck_number":    ["TRUCK NO", "TRUCK NUMBER", "VEHICLE NO"],
    "sat_slip_no":     ["SAT SLIP NO", "SATISFACTION SLIP", "SAT SLIP", "SATISFACTION SLIP NO"],
    "detention_days":  ["DETENTN DAYS", "DETENTION DAYS", "DET DAYS", "DETENTION"],
    "truck_type":      ["TRUCK TYPE", "TRPT TYPE", "VEHICLE TYPE"],
    "shipment_no":     ["SHIPMENT NO", "SHIPMENT NUMBER"],
    "service_po":      ["SERVICE PO", "SERVICE PO NO"],
    "entry_sheet_no":  ["ENTRY SHEET NO", "ENTRY SHEET"],
    "region":          ["REGION"],
    "from_location":   ["FROM", "FROM LOCATION", "ORIGIN"],
    "to_location":     ["TO", "TO LOCATION", "DESTINATION"],
    "total_units":     ["TOTAL UNITS", "UNITS"],
    "total_wt":        ["TOTAL WT", "TOTAL WEIGHT", "WEIGHT"],
    "shortage":        ["SHORTAGE", "SHORT QTY"],
    "detention_charge":["DETENTN CHARGE", "DETENTION CHARGE", "DET CHARGE"],
    "master_rate":     ["MASTER RATE", "RATE"],
    "payable":         ["PAYABLE", "AMOUNT PAYABLE", "NET PAYABLE"],
    "remarks":         ["REMARKS", "REMARK", "NOTES"],
}


def _normalize(text: str) -> str:
    return text.strip().upper().replace("_", " ") if text else ""


def _detect_header_row(ws) -> Optional[int]:
    """Find the row that contains the column headers."""
    for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        row_vals = [_normalize(str(c)) if c else "" for c in row]
        # A header row must contain at least "LR" somewhere
        if any("LR" in v for v in row_vals) and any("TRUCK" in v for v in row_vals):
            return row_idx
    return None


def _build_column_index(ws, header_row: int) -> dict:
    """Map canonical field names → 0-based column indices."""
    header_vals = [
        _normalize(str(c.value)) if c.value else ""
        for c in ws[header_row]
    ]
    index = {}
    for canonical, candidates in COLUMN_MAP.items():
        for candidate in candidates:
            try:
                col_idx = header_vals.index(candidate)
                index[canonical] = col_idx
                break
            except ValueError:
                # Try partial match
                for i, h in enumerate(header_vals):
                    if candidate in h or h in candidate:
                        index[canonical] = i
                        break
    return index


def _detect_data_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    """Return the worksheet that most likely contains the billing data."""
    for name in wb.sheetnames:
        if "bill" in name.lower():
            return wb[name]
    # Fallback: sheet with most filled rows
    best = max(wb.worksheets, key=lambda ws: ws.max_row or 0)
    return best


def _to_int(value) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value).replace(",", "").strip()))
    except (ValueError, TypeError):
        return None


def _to_float(value) -> Optional[float]:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _to_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _normalize_lr(lr_raw: str) -> str:
    """Normalize LR number: uppercase, strip spaces."""
    return lr_raw.strip().upper().replace(" ", "")


# ---------------------------------------------------------------------------
# Filename / path metadata extraction
# ---------------------------------------------------------------------------

def _extract_transporter_name(filename: str) -> str:
    """e.g., 'SENTHIL TRANSPORTS MAR'26-02.xlsx' → 'SENTHIL TRANSPORTS'"""
    stem = Path(filename).stem  # without extension
    # Pattern: NAME followed by MON'YY
    m = re.match(r"^(.+?)\s+(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)['\s]\d{2}", stem, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()
    return stem.upper()


def _extract_billing_period(filename: str) -> str:
    """e.g., 'SENTHIL TRANSPORTS MAR'26-02.xlsx' → 'MAR 2026'"""
    stem = Path(filename).stem
    m = re.search(r"(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)['\s](\d{2})", stem, re.IGNORECASE)
    if m:
        month = m.group(1).upper()
        year = int(m.group(2))
        full_year = 2000 + year if year < 100 else year
        return f"{month} {full_year}"
    return "UNKNOWN"


def _extract_client_name(file_path: str) -> str:
    """
    Extract client from folder path.
    e.g., '.../Britannia Industries Ltd/Invoice Reference/...' → 'Britannia Industries Ltd'
    """
    parts = Path(file_path).parts
    # Look for known client folders
    clients = ["Britannia Industries", "KPR Mill", "ITC Limited", "HUL"]
    for part in parts:
        for c in clients:
            if c.lower() in part.lower():
                return part
    # Fallback: parent of "Invoice Reference"
    for i, part in enumerate(parts):
        if "Invoice Reference" in part and i > 0:
            return parts[i - 1]
    return "UNKNOWN"


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_invoice_excel(file_path: str) -> InvoiceParseResult:
    """
    Parse a freight billing Excel file and return structured line items.

    Args:
        file_path: Absolute path to the .xlsx / .xls file.

    Returns:
        InvoiceParseResult with all extracted LR line items.
    """
    errors: List[str] = []
    file_path = str(file_path)
    filename = Path(file_path).name

    transporter_name = _extract_transporter_name(filename)
    billing_period = _extract_billing_period(filename)
    client_name = _extract_client_name(file_path)

    logger.info(f"Parsing Excel: {filename} | transporter={transporter_name} | period={billing_period}")

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        logger.error(f"Failed to open workbook: {exc}")
        raise

    ws = _detect_data_sheet(wb)
    sheet_name = ws.title
    logger.info(f"Using sheet: {sheet_name} (rows≈{ws.max_row})")

    header_row = _detect_header_row(ws)
    if header_row is None:
        errors.append("Could not detect header row — no LR NO / TRUCK NO headers found.")
        return InvoiceParseResult(
            transporter_name=transporter_name,
            billing_period=billing_period,
            client_name=client_name,
            total_rows=0,
            rows_needing_ocr=0,
            line_items=[],
            file_path=file_path,
            parsed_at=datetime.utcnow(),
            sheet_name=sheet_name,
            errors=errors,
        )

    col_index = _build_column_index(ws, header_row)
    logger.info(f"Column map: {col_index}")

    if "lr_number" not in col_index:
        errors.append("LR NO column not found in header row.")
        return InvoiceParseResult(
            transporter_name=transporter_name,
            billing_period=billing_period,
            client_name=client_name,
            total_rows=0,
            rows_needing_ocr=0,
            line_items=[],
            file_path=file_path,
            parsed_at=datetime.utcnow(),
            sheet_name=sheet_name,
            errors=errors,
        )

    line_items: List[InvoiceLineItem] = []

    def get_cell(row_values, key):
        idx = col_index.get(key)
        if idx is None or idx >= len(row_values):
            return None
        return row_values[idx]

    for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        # Skip fully empty rows
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        lr_raw = get_cell(row, "lr_number")
        if not lr_raw or str(lr_raw).strip() == "":
            continue  # Skip rows without an LR number

        lr_number = _normalize_lr(str(lr_raw))
        truck_type = _to_str(get_cell(row, "truck_type"))
        detention_days = _to_int(get_cell(row, "detention_days"))

        needs_ocr = truck_type is None or truck_type == ""
        needs_ocr_verify = detention_days is None

        item = InvoiceLineItem(
            lr_number=lr_number,
            truck_number=_to_str(get_cell(row, "truck_number")),
            sat_slip_no=_to_str(get_cell(row, "sat_slip_no")),
            detention_days=detention_days,
            truck_type=truck_type,
            shipment_no=_to_str(get_cell(row, "shipment_no")),
            service_po=_to_str(get_cell(row, "service_po")),
            entry_sheet_no=_to_str(get_cell(row, "entry_sheet_no")),
            region=_to_str(get_cell(row, "region")),
            from_location=_to_str(get_cell(row, "from_location")),
            to_location=_to_str(get_cell(row, "to_location")),
            total_units=_to_float(get_cell(row, "total_units")),
            total_wt=_to_float(get_cell(row, "total_wt")),
            shortage=_to_float(get_cell(row, "shortage")),
            detention_charge=_to_float(get_cell(row, "detention_charge")),
            master_rate=_to_float(get_cell(row, "master_rate")),
            payable=_to_float(get_cell(row, "payable")),
            remarks=_to_str(get_cell(row, "remarks")),
            needs_ocr=needs_ocr,
            needs_ocr_verify=needs_ocr_verify,
            row_number=row_num,
        )
        line_items.append(item)

    wb.close()

    rows_needing_ocr = sum(1 for i in line_items if i.needs_ocr)
    logger.info(
        f"Parsed {len(line_items)} LR rows | needs_ocr={rows_needing_ocr} | errors={len(errors)}"
    )

    return InvoiceParseResult(
        transporter_name=transporter_name,
        billing_period=billing_period,
        client_name=client_name,
        total_rows=len(line_items),
        rows_needing_ocr=rows_needing_ocr,
        line_items=line_items,
        file_path=file_path,
        parsed_at=datetime.utcnow(),
        sheet_name=sheet_name,
        errors=errors,
    )
