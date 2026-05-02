"""
Excel Write-back Service — IFIAS Phase 6
Writes confirmed OCR-extracted data back into the original billing Excel file.

Preserves all existing formatting, formulas, and data.
Only modifies TRUCK TYPE, DETENTN DAYS, and SAT SLIP NO cells.
Colors auto-filled cells light green for visual identification.

Usage:
    service = ExcelWritebackService()
    result = service.write_confirmed_data(
        source_excel_path="/path/to/SENTHIL TRANSPORTS MAR'26-02.xlsx",
        confirmed_items=confirmed_list,
    )
"""

import logging
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import PatternFill

from app.services.excel_parser_service import _normalize_lr, _detect_data_sheet, _detect_header_row, _build_column_index, COLUMN_MAP

logger = logging.getLogger(__name__)

# Light green fill for auto-filled cells (Excel "Good" style green)
AUTO_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
# Light yellow for manually corrected cells
MANUAL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class ConfirmedLineItem:
    lr_number: str
    truck_type: Optional[str]
    detention_days: Optional[int]
    sat_slip_no: Optional[str]
    manually_corrected: bool = False


@dataclass
class WritebackResult:
    success: bool
    cells_written: int
    rows_updated: int
    backup_path: str
    output_path: str
    errors: List[str]


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------

class ExcelWritebackService:
    """Write confirmed IFIAS data back to the original Excel billing file."""

    # Columns to write back
    WRITEBACK_COLUMNS = {
        "truck_type": ["TRUCK TYPE", "TRPT TYPE"],
        "detention_days": ["DETENTN DAYS", "DETENTION DAYS", "DET DAYS"],
        "sat_slip_no": ["SAT SLIP NO", "SATISFACTION SLIP NO"],
    }

    def write_confirmed_data(
        self,
        source_excel_path: str,
        confirmed_items: List[ConfirmedLineItem],
        output_path: Optional[str] = None,
    ) -> WritebackResult:
        """
        Write confirmed values back into the billing Excel.

        Args:
            source_excel_path: Path to original Excel file.
            confirmed_items: List of confirmed LR items with values to write.
            output_path: If None, overwrites the original (after backup).

        Returns:
            WritebackResult with counts and paths.
        """
        errors: List[str] = []
        cells_written = 0
        rows_updated = 0

        # 1. Create timestamped backup
        backup_path = self._create_backup(source_excel_path)
        logger.info(f"Backup created: {backup_path}")

        out_path = output_path or source_excel_path

        try:
            # 2. Open workbook preserving formatting
            wb = openpyxl.load_workbook(source_excel_path)
        except Exception as exc:
            logger.error(f"Cannot open workbook: {exc}")
            return WritebackResult(
                success=False,
                cells_written=0,
                rows_updated=0,
                backup_path=backup_path,
                output_path=out_path,
                errors=[str(exc)],
            )

        # 3. Detect data sheet and header
        ws = _detect_data_sheet(wb)
        logger.info(f"Writing to sheet: {ws.title}")

        header_row = _detect_header_row(ws)
        if header_row is None:
            return WritebackResult(
                success=False,
                cells_written=0,
                rows_updated=0,
                backup_path=backup_path,
                output_path=out_path,
                errors=["Header row not found in workbook"],
            )

        # 4. Build column index
        col_index = _build_column_index(ws, header_row)

        # Find LR column and target columns
        lr_col_idx = col_index.get("lr_number")
        if lr_col_idx is None:
            return WritebackResult(
                success=False,
                cells_written=0,
                rows_updated=0,
                backup_path=backup_path,
                output_path=out_path,
                errors=["LR NO column not found in header"],
            )

        truck_type_col = col_index.get("truck_type")
        detention_col = col_index.get("detention_days")
        sat_slip_col = col_index.get("sat_slip_no")

        # Build a lookup: normalized LR → ConfirmedLineItem
        # PART 6 — safety guard: only write confirmed items
        item_map: Dict[str, ConfirmedLineItem] = {
            _normalize_lr(item.lr_number): item
            for item in confirmed_items
            if getattr(item, "processing_status", "CONFIRMED").upper() == "CONFIRMED"
            or not hasattr(item, "processing_status")  # ConfirmedLineItem has no status field — already filtered
        }

        # 5. Iterate rows and write matching items
        for row in ws.iter_rows(min_row=header_row + 1):
            lr_cell = row[lr_col_idx] if lr_col_idx < len(row) else None
            if not lr_cell or not lr_cell.value:
                continue

            lr_key = _normalize_lr(str(lr_cell.value))
            item = item_map.get(lr_key)
            if not item:
                continue

            fill_style = MANUAL_YELLOW if item.manually_corrected else AUTO_FILL_GREEN
            row_written = False

            # Write truck_type
            if truck_type_col is not None and item.truck_type and truck_type_col < len(row):
                cell = row[truck_type_col]
                if not self._is_formula(cell):
                    cell.value = item.truck_type
                    cell.fill = fill_style
                    cells_written += 1
                    row_written = True

            # Write detention_days
            if detention_col is not None and item.detention_days is not None and detention_col < len(row):
                cell = row[detention_col]
                if not self._is_formula(cell):
                    cell.value = item.detention_days
                    cell.fill = fill_style
                    cells_written += 1
                    row_written = True

            # Write sat_slip_no
            if sat_slip_col is not None and item.sat_slip_no and sat_slip_col < len(row):
                cell = row[sat_slip_col]
                if not self._is_formula(cell):
                    cell.value = item.sat_slip_no
                    cell.fill = fill_style
                    cells_written += 1
                    row_written = True

            if row_written:
                rows_updated += 1

        # 6. Add IFIAS summary sheet
        try:
            self._add_summary_sheet(wb, confirmed_items, cells_written, rows_updated)
        except Exception as exc:
            logger.warning(f"Summary sheet creation failed (non-critical): {exc}")
            errors.append(f"Summary sheet: {exc}")

        # 7. Save
        try:
            wb.save(out_path)
            logger.info(f"Saved writeback to: {out_path} | rows={rows_updated} | cells={cells_written}")
        except Exception as exc:
            logger.error(f"Failed to save workbook: {exc}")
            return WritebackResult(
                success=False,
                cells_written=cells_written,
                rows_updated=rows_updated,
                backup_path=backup_path,
                output_path=out_path,
                errors=[str(exc)],
            )

        return WritebackResult(
            success=True,
            cells_written=cells_written,
            rows_updated=rows_updated,
            backup_path=backup_path,
            output_path=out_path,
            errors=errors,
        )

    def _is_formula(self, cell) -> bool:
        """Skip cells containing formulas to avoid breaking calculations."""
        return isinstance(cell.value, str) and cell.value.startswith("=")

    def _create_backup(self, source_path: str) -> str:
        """Create a timestamped backup of the Excel file."""
        p = Path(source_path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = p.parent / f"{p.stem}_backup_{ts}{p.suffix}"
        shutil.copy2(source_path, backup)
        return str(backup)

    def _add_summary_sheet(
        self,
        wb: openpyxl.Workbook,
        confirmed_items: List[ConfirmedLineItem],
        cells_written: int,
        rows_updated: int,
    ):
        """Add or replace 'IFIAS Summary' sheet with processing statistics."""
        SHEET_NAME = "IFIAS Summary"
        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]

        ws = wb.create_sheet(SHEET_NAME)
        now = datetime.now().strftime("%d-%b-%Y %H:%M")

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        from openpyxl.styles import Font
        header_font = Font(color="FFFFFF", bold=True)

        ws["A1"] = "IFIAS — Intelligent Freight Invoice Automation"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A3"] = "Processing Date:"
        ws["B3"] = now
        ws["A4"] = "Total LRs Confirmed:"
        ws["B4"] = rows_updated
        ws["A5"] = "Total Cells Updated:"
        ws["B5"] = cells_written
        ws["A6"] = "Manual Corrections:"
        ws["B6"] = sum(1 for i in confirmed_items if i.manually_corrected)
        ws["A7"] = "Auto-filled:"
        ws["B7"] = sum(1 for i in confirmed_items if not i.manually_corrected)

        # Table of confirmed items
        ws["A9"] = "LR Number"
        ws["B9"] = "Truck Type"
        ws["C9"] = "Detention Days"
        ws["D9"] = "SAT Slip No"
        ws["E9"] = "Source"

        for col in ["A9", "B9", "C9", "D9", "E9"]:
            ws[col].fill = header_fill
            ws[col].font = header_font

        for i, item in enumerate(confirmed_items, start=10):
            ws.cell(row=i, column=1, value=item.lr_number)
            ws.cell(row=i, column=2, value=item.truck_type)
            ws.cell(row=i, column=3, value=item.detention_days)
            ws.cell(row=i, column=4, value=item.sat_slip_no)
            ws.cell(row=i, column=5, value="Manual" if item.manually_corrected else "Auto")

        # Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 12
        